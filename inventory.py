import boto3
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timezone, timedelta

def remove_timezone(dt):
    if dt and isinstance(dt, str) and dt != 'N/A':
        # '2023-06-20T04:52:47.000+0000'에서 '+0000' 부분 제거
        dt = dt[:-5]
        dt = datetime.strptime(dt, '%Y-%m-%dT%H:%M:%S.%f')
        dt = dt.replace(tzinfo=None)
        return dt  # datetime 객체 반환
    return dt  # 문자열 그대로 반환

def get_ec2_instances():
    ec2_client = boto3.client('ec2')
    response = ec2_client.describe_instances()
    instances = response['Reservations']

    ec2_list = []
    for instance in instances:
        instance_details = instance['Instances'][0]
        instance_id = instance_details['InstanceId']
        instance_state = instance_details['State']['Name']
        instance_type = instance_details['InstanceType']
        private_ip = instance_details.get('PrivateIpAddress', 'N/A')
        public_ip = instance_details.get('PublicIpAddress', 'N/A')
        ec2_list.append((instance_id, instance_state, instance_type, private_ip, public_ip))

    return ec2_list

def get_ec2_tags(instance_id):
    ec2_resource = boto3.resource('ec2')
    instance = ec2_resource.Instance(instance_id)

    ec2_name = 'N/A'
    for tag in instance.tags:
        if tag['Key'] == 'Name':
            ec2_name = tag['Value']
            break

    return ec2_name
    
def get_ec2_auto_stop_tag(instance_id):
    ec2_resource = boto3.resource('ec2')
    instance = ec2_resource.Instance(instance_id)

    auto_stop_tag_value = 'N/A'
    for tag in instance.tags:
        if tag['Key'] == 'AUTO_STOP_ENABLE':
            auto_stop_tag_value = tag['Value']
            break

    return auto_stop_tag_value

def get_ec2_cpu_memory(instance_type):
    ec2_client = boto3.client('ec2')
    response = ec2_client.describe_instance_types(InstanceTypes=[instance_type])
    instance_info = response['InstanceTypes'][0]

    cpu_count = instance_info['VCpuInfo']['DefaultVCpus']
    memory_gb = instance_info['MemoryInfo']['SizeInMiB'] / 1024

    return cpu_count, memory_gb

def get_ec2_os(instance_id):
    ec2_client = boto3.client('ec2')
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    instance = response['Reservations'][0]['Instances'][0]

    # Platform 값이 Winodws 만 출력되어, 나머지는 그냥 Amazon Linux 로 설정 했다.. 
    # 나중에 다른 OS 도 쓰게되면 수정이 필요할듯
    platform_info = instance.get('Platform', 'Amazon Linux')
    return platform_info

def get_ebs_volume_size(instance_id):
    ec2_client = boto3.client('ec2')
    response = ec2_client.describe_volumes(Filters=[{'Name': 'attachment.instance-id', 'Values': [instance_id]}])
    ebs_volumes = response['Volumes']
    total_size_gb = 0
    for volume in ebs_volumes:
        total_size_gb += volume['Size']
    return f"{total_size_gb} GB"
    
def get_alb_list():
    alb_client = boto3.client('elbv2')
    response = alb_client.describe_load_balancers()
    albs = response['LoadBalancers']

    alb_list = []
    for alb in albs:
        alb_name = alb['LoadBalancerName']
        alb_dns = alb['DNSName']
        alb_type = alb['Type']
        alb_list.append((alb_name, alb_dns, alb_type))

    return alb_list

def get_s3_buckets():
    s3_client = boto3.client('s3')
    response = s3_client.list_buckets()
    buckets = response['Buckets']

    bucket_list = []
    for bucket in buckets:
        bucket_name = bucket['Name']
        bucket_list.append((bucket_name,))

    return bucket_list

def get_elasticache_clusters():
    elasticache_client = boto3.client('elasticache')
    response = elasticache_client.describe_cache_clusters()
    clusters = response['CacheClusters']

    cluster_list = []
    for cluster in clusters:
        cluster_id = cluster['CacheClusterId']
        cluster_engine = cluster['Engine']
        cluster_node_type = cluster['CacheNodeType']
        cluster_list.append((cluster_id, cluster_engine, cluster_node_type))

    return cluster_list

def get_vpc_list():
    ec2_client = boto3.client('ec2')
    response = ec2_client.describe_vpcs()
    vpcs = response['Vpcs']

    vpc_list = []
    for vpc in vpcs:
        vpc_id = vpc['VpcId']
        vpc_cidr = vpc['CidrBlock']
        vpc_list.append((vpc_id, vpc_cidr))

    return vpc_list

def get_subnet_list():
    ec2_client = boto3.client('ec2')
    response = ec2_client.describe_subnets()
    subnets = response['Subnets']

    subnet_list = []
    for subnet in subnets:
        subnet_id = subnet['SubnetId']
        subnet_cidr = subnet['CidrBlock']
        subnet_list.append((subnet_id, subnet_cidr))

    return subnet_list

def get_subnet_name(subnet_id):
    ec2_client = boto3.client('ec2')
    response = ec2_client.describe_subnets(SubnetIds=[subnet_id])
    subnet_details = response['Subnets'][0]
    return subnet_details.get('Tags', [{'Key': 'Name', 'Value': 'N/A'}])[0]['Value']
    
def get_iam_account_list():
    iam_client = boto3.client('iam')
    response = iam_client.list_users()
    users = response['Users']

    iam_account_list = []
    for user in users:
        user_name = user['UserName']
        password_last_used = user.get('PasswordLastUsed', 'N/A')
        mfa_devices = iam_client.list_mfa_devices(UserName=user_name)
        mfa_enabled = "Yes" if mfa_devices['MFADevices'] else "No"

        # 마지막 활동 시간 계산
        last_activity_days = get_iam_last_activity(user_name)

        iam_account_list.append((user_name, password_last_used, mfa_enabled, last_activity_days))

    return iam_account_list
    
def get_iam_last_activity(user_name):
    iam_client = boto3.client('iam')
    response = iam_client.get_user(UserName=user_name)
    create_date = response['User']['CreateDate'].replace(tzinfo=timezone.utc)  # Apply UTC timezone
    current_time = datetime.utcnow().replace(tzinfo=timezone.utc)  # Apply UTC timezone
    last_activity_time = current_time - create_date
    return last_activity_time.days

def get_lambda_functions():
    lambda_client = boto3.client('lambda')
    response = lambda_client.list_functions()
    functions = response['Functions']

    lambda_function_list = []
    for function in functions:
        function_name = function['FunctionName']
        function_description = function['Description']
        function_runtime = function['Runtime']
        last_modified = function['LastModified']
        lambda_function_list.append((function_name, function_description, function_runtime, last_modified))

    return lambda_function_list

def lambda_handler(event, context):
    wb = Workbook()

    # 첫 번째 탭 (EC2 리스트)
    ec2_sheet = wb.active
    ec2_sheet.title = "EC2"

    ec2_list = get_ec2_instances()
    ec2_sheet.append(["Name", "Private IP", "Public IP", "Instance Type", "CPU Count", "Memory (GB)", "OS", "EBS Volume Size", "AUTO_STOP_ENABLE"])

    for instance in ec2_list:
        instance_id, instance_state, instance_type, private_ip, public_ip = instance

        # EC2 인스턴스의 Name 태그 가져오기
        ec2_name = get_ec2_tags(instance_id)

        # EC2 인스턴스의 CPU 개수와 Memory 정보 가져오기
        cpu_count, memory_gb = get_ec2_cpu_memory(instance_type)

        # EC2 인스턴스의 OS 정보 가져오기
        ec2_os = get_ec2_os(instance_id)

        # EC2 인스턴스의 EBS 볼륨 사이즈 정보 가져오기
        ebs_volume_size = get_ebs_volume_size(instance_id)
        
        # EC2 인스턴스의 AUTO_STOP_ENABLE 태그 가져오기
        auto_stop_tag = get_ec2_auto_stop_tag(instance_id)

        ec2_sheet.append([ec2_name, private_ip, public_ip, instance_type, cpu_count, memory_gb, ec2_os, ebs_volume_size, auto_stop_tag])
        
    # 두 번째 탭 (ALB 리스트)
    alb_sheet = wb.create_sheet("ALB List")
    alb_sheet.append(["Name", "DNS Name", "Type"])

    alb_list = get_alb_list()
    for alb in alb_list:
        alb_name, alb_dns, alb_type = alb
        alb_sheet.append([alb_name, alb_dns, alb_type])

    # 세 번째 탭 (S3 리스트)
    s3_sheet = wb.create_sheet("S3")
    s3_sheet.append(["Bucket Name"])

    s3_buckets = get_s3_buckets()
    for bucket in s3_buckets:
        bucket_name = bucket[0]
        s3_sheet.append([bucket_name])

    # 네 번째 탭 (Elasticache 리스트)
    elasticache_sheet = wb.create_sheet("Elasticache")
    elasticache_sheet.append(["Cluster ID", "Engine", "Node Type"])

    elasticache_clusters = get_elasticache_clusters()
    for cluster in elasticache_clusters:
        cluster_id, cluster_engine, cluster_node_type = cluster
        elasticache_sheet.append([cluster_id, cluster_engine, cluster_node_type])

    # 다섯 번째 탭 (VPC 리스트)
    vpc_sheet = wb.create_sheet("VPC")
    vpc_sheet.append(["VPC ID", "CIDR Block"])

    vpc_list = get_vpc_list()
    for vpc in vpc_list:
        vpc_id, vpc_cidr = vpc
        vpc_sheet.append([vpc_id, vpc_cidr])

    # 여섯 번째 탭 (Subnet 리스트)
    subnet_sheet = wb.create_sheet("Subnet")
    subnet_sheet.append(["Subnet ID", "CIDR Block", "Subnet Name"])

    subnet_list = get_subnet_list()
    for subnet in subnet_list:
        subnet_id, subnet_cidr = subnet
        subnet_name = get_subnet_name(subnet_id)
        subnet_sheet.append([subnet_id, subnet_cidr, subnet_name])
        
    # 일곱 번째 탭 (IAM 계정 리스트)
    iam_sheet = wb.create_sheet("IAM Accounts")
    iam_sheet.append(["User Name", "Last Password Used", "MFA Enabled", "Last Activity (Days)"])

    iam_account_list = get_iam_account_list()
    for account in iam_account_list:
        user_name, password_last_used, mfa_enabled, last_activity_days = account
        password_last_used = remove_timezone(password_last_used)
        if isinstance(password_last_used, datetime):
            password_last_used = password_last_used.strftime('%Y-%m-%d %H:%M:%S')
        iam_sheet.append([user_name, password_last_used, mfa_enabled, last_activity_days])

    # 여덟 번째 탭 (Lambda 함수 리스트)
    lambda_sheet = wb.create_sheet("Lambda Functions")
    lambda_sheet.append(["Function Name", "Description", "Runtime", "Last Modified"])

    lambda_function_list = get_lambda_functions()
    for function in lambda_function_list:
        function_name, function_description, function_runtime, last_modified = function
        lambda_sheet.append([function_name, function_description, function_runtime, remove_timezone(last_modified)])

    # 공통 스타일 적용 함수
    def apply_common_style(sheet):
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))
                cell.border = thin_border

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))
                cell.border = thin_border

    # 스타일 적용
    apply_common_style(ec2_sheet)
    apply_common_style(alb_sheet)
    apply_common_style(s3_sheet)
    apply_common_style(elasticache_sheet)
    apply_common_style(vpc_sheet)
    apply_common_style(subnet_sheet)
    apply_common_style(iam_sheet)
    apply_common_style(lambda_sheet)

    # 열 너비 자동 조정
    for sheet in wb.sheetnames:
        for col in wb[sheet].columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            wb[sheet].column_dimensions[col[0].column_letter].width = adjusted_width

    file_path = "/tmp/aws_resources_list.xlsx"
    wb.save(file_path)
    
    # 엑셀 파일을 S3에 업로드 또는 다른 작업 수행
    # 예를 들어, S3에 업로드한다면:
    s3_client = boto3.client('s3')
    s3_bucket_name = 'YOUR-BUCKET-NAME'
    s3_key = 'EXCEL-FILE-NAME.xlsx'
    s3_client.upload_file(file_path, s3_bucket_name, s3_key)

    return {
        'statusCode': 200,
        'body': 'EC2 and S3 lists successfully generated and saved to Excel file.'
    }
